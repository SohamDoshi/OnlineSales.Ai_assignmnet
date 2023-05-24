import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# CSV file names
department_csv = 'Departments.csv'
employees_csv = 'Employees.csv'
salaries_csv = 'Salaries.csv'

# Load department data
department_data = pd.read_csv(department_csv)

# Load employees data
employees_data = pd.read_csv(employees_csv)

# Load salaries data
salaries_data = pd.read_csv(salaries_csv)

# Merge employees and salaries data based on employee ID
employees_salaries_data = pd.merge(employees_data, salaries_data, left_on="ID", right_on="EMP_ID").drop('EMP_ID', axis=1)

full_data_table = pd.merge(employees_salaries_data, department_data, left_on="DEPT ID", right_on="ID").drop('ID_y', axis=1)
#print(full_data_table)


# Calculate average monthly salary per department
department_salaries = full_data_table.groupby('NAME_y')['AMT (USD)'].mean()
#print(department_salaries)

# # Sort departments by average monthly salary in descending order
sorted_departments = department_salaries.sort_values(ascending=False).head(3)
#print(sorted_departments)


# Output directory
output_dir = 'Output'

# Create the output directory if it doesn't exist
if not os.path.exists(output_dir):
    os.makedirs(output_dir)


# Create a new workbook and worksheet
workbook = Workbook()
worksheet = workbook.active


# Write the report headers with bold styling
header_font = Font(bold=True)
worksheet.cell(row=1, column=1, value='DEPT_NAME').font = header_font
worksheet.cell(row=1, column=2, value='AVG_MONTHLY_SALARY (USD)').font = header_font

# Write the report headers
worksheet.cell(row=1, column=1, value='DEPT_NAME')
worksheet.cell(row=1, column=2, value='AVG_MONTHLY_SALARY (USD)')

# Write the top 3 departments to the worksheet
for i, (department_name, average_salary) in enumerate(sorted_departments.items()):
    worksheet.cell(row=i+2, column=1, value=department_name)
    worksheet.cell(row=i+2, column=2, value=average_salary)

# Save the workbook to the report file in the output directory
report_file = os.path.join(output_dir, 'report.xlsx')
workbook.save(report_file)


print(f'Report generated successfully: {report_file}')
